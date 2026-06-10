import openpyxl
import csv
from io import BytesIO
from django.utils.encoding import smart_str
from django.utils.dateparse import parse_date, parse_datetime

def normalize_header(header):
    """
    Clean up header string for matching.
    e.g. " Student ID  " -> "student_id"
    """
    return header.strip().lower().replace(" ", "_").replace("-", "_")

def detect_format(filename):
    """Simple detection based on extension."""
    if filename.endswith('.csv'):
        return 'csv'
    return 'xlsx'

def load_workbook_safe(file_obj):
    """Loads an xlsx file safely, handles in-memory files."""
    file_obj.seek(0)
    return openpyxl.load_workbook(file_obj, read_only=True, data_only=True)

def read_sheet_data(workbook, sheet_name=None):
    """Reads the first sheet or named sheet into a list of dicts."""
    if sheet_name and sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
    else:
        ws = workbook.active
    
    data = []
    headers = []
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [normalize_header(str(cell) if cell is not None else f"col_{j}") for j, cell in enumerate(row)]
        else:
            row_dict = {}
            for j, cell in enumerate(row):
                if j < len(headers):
                    # Try to convert to native types but keep string as fallback
                    val = cell
                    if isinstance(val, str):
                        val = val.strip()
                        # Try parsing dates
                        if '-' in val or '/' in val:
                            d = parse_date(val)
                            if d: val = d
                    row_dict[headers[j]] = val
            # Skip completely empty rows
            if any(row_dict.values()):
                data.append(row_dict)
    
    return headers, data

def generate_excel_file(data, headers):
    """Generates an in-memory Excel file from list of dicts."""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Write headers
    ws.append(headers)
    
    # Write rows
    for row in data:
        ws_row = []
        for h in headers:
            val = row.get(h, "")
            # Openpyxl handles dates/datetimes natively
            ws_row.append(val)
        ws.append(ws_row)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_csv_file(data, headers):
    """Generates an in-memory CSV file."""
    output = BytesIO()
    # Write BOM for Excel compatibility on Windows
    output.write(b'\xef\xbb\xbf')
    
    writer = csv.DictWriter(output, fieldnames=headers, extrasaction='ignore')
    writer.writeheader()
    for row in data:
        # Convert everything to string for CSV
        str_row = {}
        for k, v in row.items():
            if v is None:
                str_row[k] = ""
            else:
                str_row[k] = smart_str(v)
        writer.writerow(str_row)
    
    output.seek(0)
    return output
